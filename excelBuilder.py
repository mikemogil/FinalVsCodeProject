
import openpyxl

partNumber = ''
revnum = ''
def ExcelFiles(partNumber, partNew, rowsJob, rowsPart,revnum, non_Existing_ID):
    partNumber = partNumber
    revnum = revnum
    all_data = []
    jobscurrent = []
    startSeq = 1020
    
    rowsPart = [(partNumber, part[0], part[1], part[2], "***", revnum)for i, part in enumerate(rowsPart)]
  
    
    
    updated_part_new = [(partNumber,part[0], part[1], 0.01, "***", revnum) for i,part in enumerate(partNew)]
    if len(non_Existing_ID)>0:
        all_data = rowsPart + updated_part_new
    else:
        all_data = rowsPart

    ids = [row[1] for row in all_data]
    

    prb = openpyxl.Workbook()
    partRevBom = prb.active
    column_headers = ['PartNum', 'RevisionNum','MtlSeq', 'MtlPartNum', 'QtyPer', 'RelatedOperation', 'UOMCode', 'Plant', 'ECOGroupID', 'Company' ]
    partRevBom.append(column_headers)
    for row in all_data:
        partRevBom.append([row[0], revnum, startSeq, row[1], row[3], 10, 'Tool', 'MfgSys', 'mdieckman', 'JPMC'])
        startSeq += 10
    for row in partRevBom.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = "@"
    prb.save(fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PRB,{partNumber},{revnum}.xlsx')
    for eachJob in rowsJob:
        startSeq = 1020
        for job in all_data:
            jobscurrent.append([eachJob['JobNum'], startSeq, job[1], job[3],'Tool', 0, 10, 'MfgSys', 'JPMC', job[2]])
            startSeq += 10

    jbm = openpyxl.Workbook()
    jobBom = jbm.active
    column_headers = ['JobNum', 'MtlSeq','PartNum', 'QtyPer','IUM','AssemblySeq', 'RelatedOperation', 'Plant', 'Company', 'Description' ]
    jobBom.append(column_headers)
    for row in jobscurrent:
        jobBom.append(row)
    for row in jobBom.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = "@"
    jbm.save(fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\JB,{partNumber},{revnum}.xlsx')    
    return partNumber, revnum,all_data, ids

def secondexcelBuilder(workbook_pathPRB, workbook_pathJB, edited_dataPRB, edited_dataJB, newPartNum, newQtyPer, operatingjobs, partNumber, revnum, dataLength):
    workbookPRB = openpyxl.load_workbook(workbook_pathPRB)
    workbookJB = openpyxl.load_workbook(workbook_pathJB)
    sheetPRB = workbookPRB.active
    sheetJB = workbookJB.active
    sheetPRB.delete_rows(2, sheetPRB.max_row)
    sheetJB.delete_rows(2, sheetJB.max_row)
    for i, newparts in enumerate(newPartNum):
        edited_dataPRB.append(partNumber)
        edited_dataPRB.append(revnum)
        edited_dataPRB.append("mtl")
        edited_dataPRB.append(newparts)
        edited_dataPRB.append(newQtyPer[i])
        edited_dataPRB.append("10")
        edited_dataPRB.append("Tool")
        edited_dataPRB.append("MfgSys")
        edited_dataPRB.append("Mdieckman")
        edited_dataPRB.append("JPMC")

    seqnumber = len(dataLength) * 10 + 1020
    for jobnums in operatingjobs:
        for i, newparts in enumerate(newPartNum):
            edited_dataJB.append(jobnums)
            edited_dataJB.append(seqnumber)
            seqnumber += 10
            edited_dataJB.append(newparts)
            edited_dataJB.append(newQtyPer[i])
            edited_dataJB.append("Tool")
            edited_dataJB.append("0")
            edited_dataJB.append("10")
            edited_dataJB.append("MfgSys")
            edited_dataJB.append("JPMC")
            edited_dataJB.append("desc")
        seqnumber = len(dataLength) * 10 + 1020


    return edited_dataJB, edited_dataPRB, sheetPRB, sheetJB, workbookPRB, workbookJB
    
    
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
def NewPart(partNumber, partDescription, ClasssID, UnitPrice,RcvInspectionReq, ToolEdges2_C, ToolCat_c, ToolSubCat_c, ToolMfg_c,
            TDim_c, ToolDim_c, TRad_c, ToolRad_c, ToolAng_c, ToolFlute_c, TFluteLen_c, ToolFluteLen_c,
            TReach_c, ToolReach_c, TShank_c, ToolShank_c, ToolMtl_c, ToolCoat_c, ToolCoolant_c, tool_Notes_c,
            RevisionNum, RevShortDesc, RevDescription,drawNum, Approved, EffectiveDates, PartAuditChangeDescription,
            MinimumQty, MaximumQty, MinOrderQty, LeadTime, VendorNum):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H_%M_%S")
    EffectiveDate = current_datetime.strftime("%m/%d/%Y")


    df = pd.DataFrame({
        "PartNum": partNumber,
        "PartDescription": partDescription,
        "ClassID": ClasssID,
        "UnitPrice": UnitPrice,
        "RcvInspectionReq": RcvInspectionReq,
        "ToolEdges2_C": ToolEdges2_C,
        "ToolCat_c": ToolCat_c,
        "ToolSubCat_c": ToolSubCat_c,
        "ToolMfg_c": ToolMfg_c,
        "TDim_c": TDim_c,
        "ToolDim_c": ToolDim_c,
        "TRad_c": TRad_c,
        "ToolRad_c": ToolRad_c,
        "ToolAng_c": ToolAng_c,
        "ToolFlute_c": ToolFlute_c,
        "TFluteLen_c": TFluteLen_c,
        "ToolFluteLen_c": ToolFluteLen_c,
        "TReach_c": TReach_c,
        "ToolReach_c": ToolReach_c,
        "TShank_c": TShank_c,
        "ToolShank_c": ToolShank_c,
        "ToolMtl_c": ToolMtl_c,
        "ToolCoat_c": ToolCoat_c,
        "ToolCoolant_c": ToolCoolant_c,
        "ToolNotes_c": tool_Notes_c,
        "Company": ["JPMC"] * len(partNumber),
        "IUM": ["TOOL"] * len(partNumber),
        "PUM": ["EA"] * len(partNumber),
        "TypeCode": ["P"] * len(partNumber),
        "NonStock": ["0"] * len(partNumber),
        "PricePerCode": ["E"] * len(partNumber),
        "InternalPricePerCode": ["E"] * len(partNumber),
        "CostMethod": ["S"] * len(partNumber),
        "QtyBearing": ["1"] * len(partNumber),
    })

    df.to_excel(fr"J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\Part_DMT_{formatted_datetime}.xlsx", index=False)
    # *************************************************************************************************************************************************************************************************************
    
    data = {
        'PartNum': partNumber,
        'RevisionNum': RevisionNum,
        'RevShortDesc': RevShortDesc,
        'RevDescription': RevDescription,
        'drawNum': drawNum,
        'Approved': Approved,
        'EffectiveDate': [EffectiveDate for _ in partNumber],
        'AltMethod': ['' for _ in partNumber],
        'Company': ['JPMC' for _ in partNumber],
        'Plant': ['MfgSys' for _ in partNumber],
        'PartAudit#ChangeDescription': PartAuditChangeDescription,
    }

    df = pd.DataFrame(data)

    df.to_excel(fr"J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PartRev_DMT_{formatted_datetime}.xlsx", index=False)
    # *************************************************************************************************************************************************************************************************************
    data = {
        'PartNum': partNumber,
        'MinimumQty': MinimumQty,
        'MaximumQty': MaximumQty,
        'MinOrderQty': MinOrderQty,
        'LeadTime': LeadTime,
        'VendorNum': VendorNum,
        'PrimWhse': ['JPMC' for _ in partNumber],
        'ProcessMRP': [0 for _ in partNumber],
        'SourceType': [1 for _ in partNumber],
        'NonStock': ['P' for _ in partNumber],
        'BuyerID': [0 for _ in partNumber],
        'CostMethod': ['bpietran' for _ in partNumber],
        'QtyBearing': ['S' for _ in partNumber],
        'AutoConsumeStock': [1 for _ in partNumber],
        'RawMaterial': [1 for _ in partNumber],
        'Company': ['JPMC' for _ in partNumber],
        'Plant': ['MfgSys' for _ in partNumber],
        'PartAudit#ChangeDescription': [1 for _ in partNumber],
    }

    df = pd.DataFrame(data)

    df.to_excel(fr"J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PartPlant_DMT_{formatted_datetime}.xlsx", index=False)
    # *************************************************************************************************************************************************************************************************************
    data = {
        'Company': ['JPMC' for _ in partNumber],
        'PartNum': partNumber,
        'WarehouseCode': ['JPMC' for _ in partNumber],
        'DefaultWhse': [1 for _ in partNumber],
        'PrimBinNum': ['JPMC' for _ in partNumber],
        'Plant': ['MfgSys' for _ in partNumber],
    }

    df = pd.DataFrame(data)

    df.to_excel(fr"J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PartWhse1_DMT_{formatted_datetime}.xlsx", index=False)
    # *************************************************************************************************************************************************************************************************************
    data = {
        'Company': ['JPMC' for _ in partNumber],
        'PartNum': partNumber,
        'WarehouseCode': ['TOOLCRIB' for _ in partNumber],
        'DefaultWhse': [1 for _ in partNumber],
        'PrimBinNum': ['JPMC' for _ in partNumber],
        'Plant': ['MfgSys' for _ in partNumber],
    }

    df = pd.DataFrame(data)

    df.to_excel(fr"J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PartWhse2_DMT_{formatted_datetime}.xlsx", index=False)
    # *************************************************************************************************************************************************************************************************************
    data = {
        'Company': ['JPMC' for _ in partNumber],
        'PartNum': partNumber,
        'WarehouseCode': ['JPMC' for _ in partNumber],
        'BinNum': ['JPMC' for _ in partNumber],
        'Plant': ['MfgSys' for _ in partNumber],
    }

    df = pd.DataFrame(data)

    df.to_excel(fr"J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PartBininfo_DMT_{formatted_datetime}.xlsx", index=False)
    
    return partNumber, partDescription

