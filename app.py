from flask import Flask, render_template, request, redirect, url_for
import os
from data_file import part, getDescription
import openpyxl
from findfile import get_latest_file_path, id_list_from_file
from excelBuilder import NewPart
from data_file import getJobNum, getInvalids
import pandas as pd

directory_path = r"X:\PROGRAMMING\CUSTOMER"
dropdown = [folder for folder in os.listdir(directory_path) if os.path.isdir(os.path.join(directory_path, folder))]

app = Flask(__name__)
# data_file.engine

@app.route('/')
def form():
    
    return render_template('main.html', dropdown = dropdown)

# **************************************************************************************************************************************************************************************



@app.route('/submit-form', methods=['POST'])
def submit_form():

    if 'selected_files' in request.form:
        part_number = request.form.get('part_number')
        selected_files = request.form.getlist('selected_files')
        revision_number = request.form.get('revision_number')
        dropdown_value = request.form.get('dropdown_value')
        partnum_values_ID = id_list_from_file(selected_files)
        invalidIDs = getInvalids(partnum_values_ID)
        if len(invalidIDs) > 0:
          
          return redirect(url_for('invalidP', part_number=part_number, revision_number=revision_number, invalIDs=invalidIDs, all_ids=partnum_values_ID))
        elif len(invalidIDs) == 0:
          return redirect(url_for('success', part_number=part_number, revision_number=revision_number, selected_files=partnum_values_ID, dropdown_value=dropdown_value))
    else:
        part_number = request.form['partNumber']
        part_number = part_number.replace(" ", "")
        revision_number = request.form['revisionNumber']
        dropdown_value = request.form['dropdown_value']
        filepaths = get_latest_file_path(part_number, dropdown_value)
        return render_template('files.html', part_number=part_number, revision_number=revision_number, dropdown_value=dropdown_value, selected_files=filepaths[0], latest_file=filepaths[1])




# **************************************************************************************************************************************************************************************
@app.route('/create-part', methods=['POST'])
def create_part():
    return render_template('makeNewPart2.html')


@app.route('/process-part', methods=['POST'])
def process_part():
    partNumber = request.form.getlist('PartNumber')
    part_descriptions = request.form.getlist('PartDescription')
    class_ids = request.form.getlist('ClassID')
    unit_prices = request.form.getlist('UnitPrice')
    RcvInspectionReq = request.form.getlist('RcvInspectionReq')
    tool_edges = request.form.getlist('ToolEdges2_C')
    tool_cats = request.form.getlist('ToolCat_c')
    tool_Notes_c = request.form.getlist('Tool_Notes_c')
    tool_subcats = request.form.getlist('ToolSubCat_c')
    tool_mfgs = request.form.getlist('ToolMfg_c')
    tdims = request.form.getlist('TDim_c')
    tooldims = request.form.getlist('ToolDim_c')
    trads = request.form.getlist('TRad_c')
    toolrads = request.form.getlist('ToolRad_c')
    tool_angs = request.form.getlist('ToolAng_c')
    tool_flutes = request.form.getlist('ToolFlute_c')
    tflute_lens = request.form.getlist('TFluteLen_c')
    tool_flute_lens = request.form.getlist('ToolFluteLen_c')
    treaches = request.form.getlist('TReach_c')
    tool_reaches = request.form.getlist('ToolReach_c')
    tshanks = request.form.getlist('TShank_c')
    tool_shanks = request.form.getlist('ToolShank_c')
    tool_mtls = request.form.getlist('ToolMtl_c')
    tool_coats = request.form.getlist('ToolCoat_c')
    tool_coolants = request.form.getlist('ToolCoolant_c')
    revision_nums = request.form.getlist('RevisionNum')
    rev_short_descs = request.form.getlist('RevShortDesc')
    rev_descs = request.form.getlist('RevDescription')
    drawNum = request.form.getlist('DrawNum')
    approved = request.form.getlist('Approved')
    effective_dates = request.form.getlist('EffectiveDate')
    part_audit_descs = request.form.getlist('PartAudit#ChangeDescription')
    min_qtys = request.form.getlist('MinimumQty')
    max_qtys = request.form.getlist('MaximumQty')
    min_order_qtys = request.form.getlist('MinOrderQty')
    lead_times = request.form.getlist('LeadTime')
    vendor_nums = request.form.getlist('VendorNum')
    print(part_descriptions)
    NewlyCreatedParts = NewPart(partNumber, part_descriptions, class_ids, unit_prices, RcvInspectionReq, tool_edges, tool_cats, tool_subcats, tool_mfgs, tdims, tooldims, trads, toolrads, tool_angs, tool_flutes, tflute_lens, tool_flute_lens, treaches, tool_reaches, tshanks, tool_shanks, tool_mtls, tool_coats, tool_coolants, tool_Notes_c, revision_nums, rev_short_descs, rev_descs,drawNum, approved, effective_dates, part_audit_descs, min_qtys, max_qtys, min_order_qtys, lead_times, vendor_nums)
    bob = NewlyCreatedParts[0]
    print(bob)
    
    return render_template('finished.html')
# **************************************************************************************************************************************************************************************
@app.route('/invalidP', methods=['GET', 'POST'])
def invalidP():
     
     if request.method == 'POST':
         if 'skip' in request.form:
            part_number = request.form.get('part_number')
            revision_number = request.form.get('revision_number')
            partnum_values_ID = request.form.get('all_ids')
            partnum_values_ID = eval(partnum_values_ID)
            return redirect(url_for('success', part_number=part_number, revision_number=revision_number, selected_files=partnum_values_ID))
         if 'process' in request.form:
            part_number = request.form.get('part_number')
            revision_number = request.form.get('revision_number')
            partnum_values_ID = request.form.get('all_ids')
            partnum_values_ID = eval(partnum_values_ID)
            invalidIDs = request.form.getlist('invalIDs')
            checked_values = request.form.getlist('checkedValues')
            for value in checked_values:
                valid = getDescription(value)
                
                if len(valid) > 0:
                    partnum_values_ID.append(value)
                    checked_values.remove(value)

            return render_template('makeNewPart.html', checked_values=checked_values, partnum_values_ID = partnum_values_ID, part_number = part_number, revision_number=revision_number)
            
     else:
        part_number = request.args.get('part_number')
        revision_number = request.args.get('revision_number')
        all_values_ID = request.args.getlist('all_ids')
        invalidIDs = request.args.getlist('invalIDs')
        
        return render_template('invalidparts.html', part_number=part_number, revision_number=revision_number, all_ids=all_values_ID, invalIDs=invalidIDs)
     if "ClassID" in request.form:
        revision_number = request.form.get('revision_number')
        checked_values = request.form.get('checked_values')
        partnum_values_ID = request.form.get('partnum_values_ID')
        checked_values = eval(checked_values)
        partnum_values_ID = eval(partnum_values_ID)
        part_number = request.form.get('part_number')
        part_descriptions = request.form.getlist('PartDescription')
        class_ids = request.form.getlist('ClassID')
        unit_prices = request.form.getlist('UnitPrice')
        RcvInspectionReq = request.form.getlist('RcvInspectionReq')
        tool_edges = request.form.getlist('ToolEdges2_C')
        tool_cats = request.form.getlist('ToolCat_c')
        tool_Notes_c = request.form.getlist('Tool_Notes_c')
        tool_subcats = request.form.getlist('ToolSubCat_c')
        tool_mfgs = request.form.getlist('ToolMfg_c')
        tdims = request.form.getlist('TDim_c')
        tooldims = request.form.getlist('ToolDim_c')
        trads = request.form.getlist('TRad_c')
        toolrads = request.form.getlist('ToolRad_c')
        tool_angs = request.form.getlist('ToolAng_c')
        tool_flutes = request.form.getlist('ToolFlute_c')
        tflute_lens = request.form.getlist('TFluteLen_c')
        tool_flute_lens = request.form.getlist('ToolFluteLen_c')
        treaches = request.form.getlist('TReach_c')
        tool_reaches = request.form.getlist('ToolReach_c')
        tshanks = request.form.getlist('TShank_c')
        tool_shanks = request.form.getlist('ToolShank_c')
        tool_mtls = request.form.getlist('ToolMtl_c')
        tool_coats = request.form.getlist('ToolCoat_c')
        tool_coolants = request.form.getlist('ToolCoolant_c')
        revision_nums = request.form.getlist('RevisionNum')
        rev_short_descs = request.form.getlist('RevShortDesc')
        rev_descs = request.form.getlist('RevDescription')
        drawNum = request.form.getlist('DrawNum')
        approved = request.form.getlist('Approved')
        effective_dates = request.form.getlist('EffectiveDate')
        part_audit_descs = request.form.getlist('PartAudit#ChangeDescription')
        min_qtys = request.form.getlist('MinimumQty')
        max_qtys = request.form.getlist('MaximumQty')
        min_order_qtys = request.form.getlist('MinOrderQty')
        lead_times = request.form.getlist('LeadTime')
        vendor_nums = request.form.getlist('VendorNum')
        
        NewlyCreatedParts = NewPart(checked_values, part_descriptions, class_ids, unit_prices, RcvInspectionReq, tool_edges, tool_cats, tool_subcats, tool_mfgs, tdims, tooldims, trads, toolrads, tool_angs, tool_flutes, tflute_lens, tool_flute_lens, treaches, tool_reaches, tshanks, tool_shanks, tool_mtls, tool_coats, tool_coolants, tool_Notes_c, revision_nums, rev_short_descs, rev_descs,drawNum, approved, effective_dates, part_audit_descs, min_qtys, max_qtys, min_order_qtys, lead_times, vendor_nums)
        return redirect(url_for('success',checked_values = checked_values, part_number=part_number, revision_number=revision_number, selected_files=partnum_values_ID, NewParts = NewlyCreatedParts[0], NewPartsDesc = NewlyCreatedParts[1]))

# **************************************************************************************************************************************************************************************
import pandas as pd


@app.route('/success', methods=['GET', 'POST'])
def success():
    if request.method == 'POST':
        # Get the edited data from the form
        selected_files = request.form.get('selected_files')
        
        selected_files = eval(selected_files)
        part_number = request.form.get('part_number')
        checked_values = request.form.get('checked_values') 
        if checked_values == None:
            checked_values = []
        else:
            checked_values = eval(checked_values)
                
        revision_number = request.form.get('revision_number')
        edited_dataPRB = request.form.getlist('edited_dataPRB')
        edited_dataJB = request.form.getlist('edited_dataJB')
        newPartNum = request.form.getlist('partNum')
        newQtyPer = request.form.getlist('qtyPer')
        dataLength = request.form.getlist('jobs')
        selected_files = request.form.get('selected_files')
        jobs = getJobNum(part_number, revision_number)
        selected_files = eval(selected_files)
        print(len(selected_files),"jobs")
        # part(part_number, revision_number, selected_files)
        workbook_pathPRB = fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PRB,{part_number},{revision_number}.xlsx'
        workbook_pathJB = fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\JB,{part_number},{revision_number}.xlsx'
        workbookPRB = openpyxl.load_workbook(workbook_pathPRB)
        workbookJB = openpyxl.load_workbook(workbook_pathJB)
        sheetPRB = workbookPRB.active
        sheetJB = workbookJB.active
        sheetPRB.delete_rows(2, sheetPRB.max_row)
        sheetJB.delete_rows(2, sheetJB.max_row)
        for i, newparts in enumerate(newPartNum):
                edited_dataPRB.append(part_number)
                edited_dataPRB.append(revision_number)
                edited_dataPRB.append("mtl")
                edited_dataPRB.append(newparts)
                edited_dataPRB.append(newQtyPer[i])
                edited_dataPRB.append("10")
                edited_dataPRB.append("Tool")
                edited_dataPRB.append("MfgSys")
                edited_dataPRB.append("Mdieckman")
                edited_dataPRB.append("JPMC")

 

        seqnumber = len(selected_files) * 10 + 1020
        print(seqnumber, "SEQNUMBER")
        for jobnums in jobs:
            for i, newparts in enumerate(newPartNum):
                print(newparts, "HERE THEY ARE")
                edited_dataJB.append(jobnums)
                edited_dataJB.append(seqnumber)
                seqnumber += 10
                edited_dataJB.append(newparts)
                edited_dataJB.append(newQtyPer[i])
                edited_dataJB.append("Tool")
                edited_dataJB.append("0")
                edited_dataJB.append("10")
                edited_dataJB.append("MfgSys")
                edited_dataJB.append("JPMC")
                edited_dataJB.append("desc")
            seqnumber = len(selected_files) * 10 + 1020
        data = {
        'PartNum': [],
        'RevisionNum': [],
        'MtlSeq': [],
        'MtlPartNum': [],
        'QtyPer': [],
        'RelatedOperation': [],
        'UOMCode': [],
        'Plant': [],
        'ECOGroupID': [],
        'Company': []}
        data2 = {
        'JobNum': [],
        'MtlSeq': [],
        'PartNum': [],
        'QtyPer': [],
        'IUM': [],
        'AssemblySeq': [],
        'RelatedOperation': [],
        'Plant': [],
        'Company': [],
        'Description': []}
        seqPRB = 1020
        mapJB = {}
        array1 = []
        array2 = []
        for i, value in enumerate(edited_dataPRB):

            row_number = i // sheetPRB.max_column + 2  # Add 2 to skip the header row
            col_number = i % sheetPRB.max_column + 1
            if col_number == 1:
                cellvalueP = str(part_number)
                data['PartNum'].append(cellvalueP)
            elif col_number == 2:
                cellvalueP = revision_number
                b = str(cellvalueP)
                data['RevisionNum'].append(cellvalueP)
            elif col_number == 3:
                cellvalueP = seqPRB
                c = str(cellvalueP)
                data['MtlSeq'].append(cellvalueP)
            elif col_number == 4:
                cellvalueP = value
                mapJB[value] = seqPRB
                seqPRB += 10       
                d = str(cellvalueP)
                data['MtlPartNum'].append(cellvalueP)
            elif col_number == 5:
                cellvalueP = value
                e = str(cellvalueP)
                data['QtyPer'].append(cellvalueP)
            elif col_number == 6:
                cellvalueP = '10'
                f = str(cellvalueP)
                data['RelatedOperation'].append(cellvalueP)
            elif col_number == 7:
                cellvalueP = 'Tool'
                g = str(cellvalueP)
                data['UOMCode'].append(cellvalueP)
            elif col_number == 8:
                cellvalueP = 'MfgSys'
                h = str(cellvalueP)
                data['Plant'].append(cellvalueP)
            elif col_number == 9:
                cellvalueP = 'mdieckman'
                p = 'mdieckman'
                data['ECOGroupID'].append(cellvalueP)
            elif col_number == 10:
                cellvalueP = "JPMC"
                j = str(cellvalueP)
                data['Company'].append(cellvalueP)
            else:
                cellvalueP = str(cellvalueP)

            sheetPRB.cell(row=row_number, column=col_number).value = cellvalueP
            

        counter = 0
        
        for i, value in enumerate(edited_dataJB):
            row_number = i // sheetJB.max_column + 2  # Add 2 to skip the header row
            col_number = i % sheetJB.max_column + 1
            if col_number == 1:
                cellvalueJ = value
                data2['JobNum'].append(cellvalueJ)
            elif col_number == 2:
                cellvalueJ = value
                data2['MtlSeq'].append(cellvalueJ)
            elif col_number == 3:
                cellvalueJ = value
                data2['PartNum'].append(cellvalueJ)
            elif col_number == 4:
                data2['QtyPer'].append(value)
            elif col_number == 5:
                cellvalueJ = 'Tool'
                data2['IUM'].append(cellvalueJ)
            elif col_number == 6:
                cellvalueJ = '0'
                data2['AssemblySeq'].append(cellvalueJ)
            elif col_number == 7:
                cellvalueJ = '10'
                data2['RelatedOperation'].append(cellvalueJ)
            elif col_number == 8:
                cellvalueJ = 'MfgSys'
                data2['Plant'].append(cellvalueJ)   
            elif col_number == 9:
                cellvalueJ = 'JPMC'
                data2['Company'].append(cellvalueJ)
            elif col_number == 10:
                if (sheetJB.cell(row=row_number, column=3).value) in checked_values:
                    cellvalueJ = value
                    data2['Description'].append(cellvalueJ)
                    
                else:
                    cellvalueJ = str(getDescription(sheetJB.cell(row=row_number, column=3).value))[2:-3]
                    data2['Description'].append(cellvalueJ)

                
                
            else:
                cellvalueJ = value

            # print(sheetJB.cell(row=row_number, column=3).value, value)
            counter += 1
            sheetJB.cell(row=row_number, column=col_number).value = cellvalueJ

        # Save the changes back to the Excel file
        for row in sheetPRB.iter_rows():
            for cell in row:
                cell.number_format = '@'


        for row in sheetJB.iter_rows():
            for cell in row:
                cell.number_format = '@'


        workbookPRB.save(workbook_pathPRB)
        workbookJB.save(workbook_pathJB)
        os.remove(fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PRB,{part_number},{revision_number}.xlsx')
        os.remove(fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\JB,{part_number},{revision_number}.xlsx')

        

        df1 = pd.DataFrame(data)
        df1.to_excel(fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PRB,{part_number},{revision_number}.xlsx', index=False)
        df2 = pd.DataFrame(data2)
        df2.to_excel(fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\JB,{part_number},{revision_number}.xlsx', index=False)

        return render_template('finished.html')
    else:
        # Get the initial data from the Excel sheet
        selected_files = request.args.getlist('selected_files')
        dropdown = request.args.get('dropdown_value')
        newParts = request.args.getlist('NewParts')
        newPartsDesc = request.args.getlist('NewPartsDesc')
        checked_values = request.args.getlist('checked_values')
        print(checked_values, "Checked")
        part_number = request.args.get('part_number')
        revision_number = request.args.get('revision_number')

        partinfo = part(part_number, revision_number, selected_files, newParts, newPartsDesc)
        workbook_pathPRB = fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\PRB,{part_number},{revision_number}.xlsx'
        workbook_pathJB = fr'J:\ERP-Business Intelligence\Bill of Materials (BOM)\BOM output\JB,{part_number},{revision_number}.xlsx'

        workbookPRB = openpyxl.load_workbook(workbook_pathPRB)
        workbookJB = openpyxl.load_workbook(workbook_pathJB)
        sheetPRB = workbookPRB.active
        sheetJB = workbookJB.active

        headersPRB = [cell.value for cell in sheetPRB[1]]
        headersJB = [cell.value for cell in sheetJB[1]]
        dataPRB = []
        dataJB = []

        for rowPRB in sheetPRB.iter_rows(min_row=2, values_only=True):
            dataPRB.append(rowPRB)

        for rowJB in sheetJB.iter_rows(min_row=2, values_only=True):
            dataJB.append(rowJB)

        return render_template('excel.html', part_number=part_number, revision_number=revision_number,
                               headersPRB=headersPRB, dataPRB=dataPRB, dataJB=dataJB, headersJB=headersJB,
                               selected_files=partinfo[6], jobs=partinfo[4], checked_values = checked_values)
    
# **************************************************************************************************************************************************************************************


if __name__ == '__main__':
    app.run(debug=True)

# , host='phl-ws-0025', port=8050
